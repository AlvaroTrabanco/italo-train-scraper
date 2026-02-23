#!/usr/bin/env python3
import argparse
import csv
import os
import re
import sys
import zipfile
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

from openpyxl import load_workbook, Workbook


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def read_gtfs_routes_from_zip(zip_path: str) -> List[Dict[str, str]]:
    """
    Reads routes.txt from the GTFS zip and returns rows as dicts.
    Expects: route_id, agency_id, route_short_name, route_long_name, route_type
    """
    if not os.path.exists(zip_path):
        raise SystemExit(f"GTFS zip not found: {zip_path}")

    with zipfile.ZipFile(zip_path, "r") as z:
        if "routes.txt" not in z.namelist():
            raise SystemExit(f"routes.txt not found inside: {zip_path}")
        raw = z.read("routes.txt").decode("utf-8", "replace").splitlines()

    if not raw:
        return []

    header = [h.strip() for h in raw[0].split(",")]
    rows: List[Dict[str, str]] = []
    for line in raw[1:]:
        # routes.txt here is simple CSV without quotes/commas in fields; safe split
        parts = line.split(",")
        if len(parts) < len(header):
            parts += [""] * (len(header) - len(parts))
        d = {header[i]: (parts[i] if i < len(parts) else "") for i in range(len(header))}
        rows.append(d)
    return rows


def parse_route_long_name(route_long_name: str) -> Tuple[str, str]:
    """
    Your build_gtfs writes route_long_name as "{origin} – {dest}" (note EN DASH).
    Fallbacks included.
    """
    s = (route_long_name or "").strip()
    if " – " in s:
        a, b = s.split(" – ", 1)
        return a.strip(), b.strip()
    if " - " in s:
        a, b = s.split(" - ", 1)
        return a.strip(), b.strip()
    # last resort: cannot parse
    return s.strip(), ""


def load_expected_xlsx(path: str) -> List[Dict[str, str]]:
    if not os.path.exists(path):
        raise SystemExit(f"Expected routes xlsx not found: {path}")

    wb = load_workbook(path)
    ws = wb.active

    # Header row is row 1
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    header_idx = {h: i for i, h in enumerate(headers) if h}

    # We’ll try common column names without forcing you to rename anything.
    # Priority: mapped names if present, otherwise raw.
    dep_keys = ["Departure_mapped", "From_mapped", "Departure", "From"]
    arr_keys = ["Arrival_mapped", "To_mapped", "Arrival", "To"]

    def pick_key(keys: List[str]) -> Optional[str]:
        for k in keys:
            if k in header_idx:
                return k
        return None

    dep_col = pick_key(dep_keys)
    arr_col = pick_key(arr_keys)

    if not dep_col or not arr_col:
        raise SystemExit(
            f"Expected file must contain Departure/Arrival columns (or *_mapped). "
            f"Found headers: {headers}"
        )

    out: List[Dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        dep = ws.cell(row=r, column=header_idx[dep_col] + 1).value
        arr = ws.cell(row=r, column=header_idx[arr_col] + 1).value
        dep_s = (str(dep).strip() if dep is not None else "")
        arr_s = (str(arr).strip() if arr is not None else "")
        if not dep_s and not arr_s:
            continue
        out.append({"departure": dep_s, "arrival": arr_s})

    return out


@dataclass
class MatchRow:
    departure: str
    arrival: str
    status: str
    trains: str
    gtfs_route_long_names: str


def main() -> None:
    ap = argparse.ArgumentParser(description="Compare expected A→B routes against GTFS routes.txt and output missing list.")
    ap.add_argument("--expected-xlsx", required=True, help="Path to expected_routes_mapped.xlsx")
    ap.add_argument("--gtfs-zip", required=True, help="Path to GTFS zip (italo_latest.zip or dated zip)")
    ap.add_argument("--out-dir", required=True, help="Output directory (e.g. public/reports)")
    ap.add_argument("--out-prefix", default="missing_routes", help="Output basename prefix (default missing_routes)")
    args = ap.parse_args()

    os.makedirs(args.out_dir, exist_ok=True)

    expected = load_expected_xlsx(args.expected_xlsx)
    gtfs_routes = read_gtfs_routes_from_zip(args.gtfs_zip)

    # Index GTFS by normalized (origin,dest)
    idx: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
    for r in gtfs_routes:
        long_name = r.get("route_long_name", "") or ""
        short = r.get("route_short_name", "") or ""  # train number in your feed
        a, b = parse_route_long_name(long_name)
        key = (norm(a), norm(b))
        idx.setdefault(key, []).append({"train": short.strip(), "long_name": long_name.strip()})

    results: List[MatchRow] = []
    missing_count = 0

    for e in expected:
        dep = e["departure"]
        arr = e["arrival"]
        key = (norm(dep), norm(arr))
        matches = idx.get(key, [])

        if matches:
            trains = sorted({m["train"] for m in matches if m["train"]})
            long_names = sorted({m["long_name"] for m in matches if m["long_name"]})
            results.append(
                MatchRow(
                    departure=dep,
                    arrival=arr,
                    status="OK",
                    trains=", ".join(trains),
                    gtfs_route_long_names=" | ".join(long_names),
                )
            )
        else:
            missing_count += 1
            results.append(
                MatchRow(
                    departure=dep,
                    arrival=arr,
                    status="MISSING_IN_GTFS",
                    trains="",
                    gtfs_route_long_names="",
                )
            )

    # Write CSV
    csv_path = os.path.join(args.out_dir, f"{args.out_prefix}_latest.csv")
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["departure", "arrival", "status", "trains", "gtfs_route_long_names"])
        for r in results:
            w.writerow([r.departure, r.arrival, r.status, r.trains, r.gtfs_route_long_names])

    # Write Markdown (easy visual scan in Pages)
    md_path = os.path.join(args.out_dir, f"{args.out_prefix}_latest.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# Missing routes report\n\n")
        f.write(f"- Expected pairs: **{len(results)}**\n")
        f.write(f"- Missing in GTFS: **{missing_count}**\n\n")

        f.write("## Missing (Expected but not in GTFS)\n\n")
        f.write("| departure | arrival |\n|---|---|\n")
        for r in results:
            if r.status == "MISSING_IN_GTFS":
                f.write(f"| {r.departure} | {r.arrival} |\n")

        f.write("\n## Present (Expected and found in GTFS)\n\n")
        f.write("| departure | arrival | trains |\n|---|---|---|\n")
        for r in results:
            if r.status == "OK":
                f.write(f"| {r.departure} | {r.arrival} | {r.trains} |\n")

        f.write("\n")

    # Write XLSX (useful for sorting/filtering)
    xlsx_path = os.path.join(args.out_dir, f"{args.out_prefix}_latest.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "missing_routes"
    ws.append(["departure", "arrival", "status", "trains", "gtfs_route_long_names"])
    for r in results:
        ws.append([r.departure, r.arrival, r.status, r.trains, r.gtfs_route_long_names])
    wb.save(xlsx_path)

    print(f"Wrote:\n- {csv_path}\n- {md_path}\n- {xlsx_path}\nMissing count: {missing_count}")


if __name__ == "__main__":
    main()